import os
import tempfile
from datetime import timedelta
from unittest.mock import MagicMock, patch

from django.contrib.auth.models import User
from django.test import override_settings
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APITestCase, APITransactionTestCase

from autenticacion.models import CorreoAutorizado
from catalogos.models import Estatus, MedioRecepcion, UnidadAdministrativa
from .models import FUS, Evidencia, Turnado, Bitacora, Notificacion, Actividad


class BitacoraROL2Tests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        for clave, tipo, orden in (
            ('Registrado', 'PARTICULAR', 1),
            ('Recibido', 'TITULAR', 2),
        ):
            Estatus.objects.get_or_create(
                clave=clave,
                defaults={'nombre': clave, 'tipoFlujo': tipo, 'orden': orden},
            )

        cls.rol1 = User.objects.create_user(
            username='dueno-bitacora@anam.gob.mx',
            email='dueno-bitacora@anam.gob.mx',
            password='x',
        )
        cls.rol2 = User.objects.create_user(
            username='titular-a@anam.gob.mx', email='titular-a@anam.gob.mx', password='x',
        )
        cls.otro_rol2 = User.objects.create_user(
            username='titular-b@anam.gob.mx', email='titular-b@anam.gob.mx', password='x',
        )
        for usuario, rol in ((cls.rol1, 'ROL1'), (cls.rol2, 'ROL2'), (cls.otro_rol2, 'ROL2')):
            CorreoAutorizado.objects.create(
                email=usuario.email,
                nombre=usuario.username,
                rol=rol,
                activo=1,
            )

        cls.fus_propio = FUS.objects.create(
            folio='FUS/BITACORA/ROL2/001',
            idSolicitanteInterno=cls.rol1,
            descripcion='FUS asignado al titular A',
            contexto='',
            estatusParticular_id='Registrado',
        )
        cls.fus_ajeno = FUS.objects.create(
            folio='FUS/BITACORA/ROL2/002',
            idSolicitanteInterno=cls.rol1,
            descripcion='FUS asignado al titular B',
            contexto='',
            estatusParticular_id='Registrado',
        )
        Turnado.objects.create(
            idFus=cls.fus_propio,
            idRemitente=cls.rol1,
            idDestinatario=cls.rol2,
            estatusTitular_id='Recibido',
        )
        Turnado.objects.create(
            idFus=cls.fus_ajeno,
            idRemitente=cls.rol1,
            idDestinatario=cls.otro_rol2,
            estatusTitular_id='Recibido',
        )

        Bitacora.objects.create(
            usuario=cls.rol2.email,
            rol='ROL2',
            accion='ASIGNACION_ESTADO',
            fusFolio=cls.fus_propio.folio,
            estadoAnterior='Turnado',
            estadoNuevo='Atendido',
        )
        Bitacora.objects.create(
            usuario=cls.rol2.email,
            rol='ROL2',
            accion='REGISTRO_RESPUESTA',
            fusFolio=cls.fus_propio.folio,
        )
        Bitacora.objects.create(
            usuario=cls.rol2.email,
            rol='ROL2',
            accion='ASIGNACION_ESTADO',
            fusFolio=cls.fus_propio.folio,
            estadoAnterior='En_seguimiento',
            estadoNuevo='Atendido',
        )
        Bitacora.objects.create(
            usuario=cls.otro_rol2.email,
            rol='ROL2',
            accion='REGISTRO_RESPUESTA',
            fusFolio=cls.fus_ajeno.folio,
        )
        Bitacora.objects.create(
            usuario=cls.rol1.email,
            rol='ROL1',
            accion='TURNAR_FUS',
            fusFolio=cls.fus_propio.folio,
            estadoAnterior='Registrado',
            estadoNuevo='Turnado',
        )

    def setUp(self):
        self.client.force_authenticate(self.rol2)

    def test_rol2_solo_recibe_sus_movimientos_de_fus_asignados(self):
        response = self.client.get('/api/bitacora/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['total'], 3)
        self.assertTrue(all(r['usuario'] == self.rol2.email for r in response.data['results']))
        self.assertTrue(all(r['fusFolio'] == self.fus_propio.folio for r in response.data['results']))

    def test_rol2_ve_estados_traducidos_a_su_flujo(self):
        response = self.client.get('/api/bitacora/')
        transiciones = {(r['estadoAnterior'], r['estadoNuevo']) for r in response.data['results']}
        self.assertIn(('Recibido', 'En_seguimiento'), transiciones)
        self.assertIn(('En_seguimiento', 'Pendiente_validacion'), transiciones)

    def test_filtro_rol2_usa_estados_del_titular(self):
        en_seguimiento = self.client.get('/api/bitacora/?estatus_fus=En_seguimiento')
        pendiente = self.client.get('/api/bitacora/?estatus_fus=Pendiente_validacion')
        self.assertEqual(en_seguimiento.data['total'], 2)
        self.assertEqual(pendiente.data['total'], 1)

    def test_excel_y_pdf_respetan_filtros_y_columnas_seleccionadas(self):
        import io
        import openpyxl

        query = '?estatus_fus=Pendiente_validacion&columnas=folio,estatus'
        excel = self.client.get(f'/api/bitacora/exportar/excel/{query}')
        self.assertEqual(excel.status_code, status.HTTP_200_OK)
        self.assertEqual(
            excel['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        libro = openpyxl.load_workbook(io.BytesIO(excel.content), read_only=True)
        hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        indice_encabezado = next(
            i for i, fila in enumerate(filas)
            if tuple(fila[:2]) == ('Folio', 'Estatus')
        )
        datos = [fila[:2] for fila in filas[indice_encabezado + 1:] if fila[0]]
        self.assertEqual(datos, [(self.fus_propio.folio, 'Pendiente de validación')])

        pdf = self.client.get(f'/api/bitacora/exportar/pdf/{query}')
        self.assertEqual(pdf.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))


class BitacoraROL1ExportTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.unidad = UnidadAdministrativa.objects.create(
            idUnidadAdministrativa=99001,
            clave='QA-BITACORA',
            unidadAdministrativa='Unidad de prueba Bitácora',
            esUnidadAdministrativa=1,
            activo=1,
        )
        cls.rol1 = User.objects.create_user(
            username='particular-bitacora@anam.gob.mx',
            email='particular-bitacora@anam.gob.mx',
            password='x',
        )
        cls.responsable = User.objects.create_user(
            username='responsable-bitacora@anam.gob.mx',
            email='responsable-bitacora@anam.gob.mx',
            password='x',
        )
        cls.otro_responsable = User.objects.create_user(
            username='otro-responsable@anam.gob.mx',
            email='otro-responsable@anam.gob.mx',
            password='x',
        )
        CorreoAutorizado.objects.create(
            email=cls.rol1.email, nombre='Particular de prueba', rol='ROL1', activo=1,
        )
        CorreoAutorizado.objects.create(
            email=cls.responsable.email,
            nombre='Responsable correcto',
            rol='ROL2',
            unidadAdministrativa=cls.unidad,
            activo=1,
        )
        CorreoAutorizado.objects.create(
            email=cls.otro_responsable.email, nombre='Responsable ajeno', rol='ROL2', activo=1,
        )
        Bitacora.objects.create(
            usuario=cls.responsable.email,
            rol='ROL2',
            accion='CONCLUSION_FUS',
            fusFolio='ANAM/PARTICULAR/FUS/ROL1/001',
            estadoAnterior='Pendiente_validacion',
            estadoNuevo='Concluido',
            observaciones='Movimiento que debe exportarse',
        )
        Bitacora.objects.create(
            usuario=cls.otro_responsable.email,
            rol='ROL2',
            accion='ASIGNACION_ESTADO',
            fusFolio='ANAM/PARTICULAR/FUS/ROL1/002',
            estadoAnterior='Turnado',
            estadoNuevo='Atendido',
            observaciones='Movimiento que debe quedar fuera',
        )

    def setUp(self):
        self.client.force_authenticate(self.rol1)

    def _query(self, columnas=None):
        from urllib.parse import urlencode
        from django.utils import timezone

        params = [
            ('q', 'FUS/ROL1/001'),
            ('usuario', self.responsable.email),
            ('unidadAdministrativa', str(self.unidad.idUnidadAdministrativa)),
            ('estatus_fus', 'Concluido'),
            ('fecha_desde', timezone.localdate().isoformat()),
            ('fecha_hasta', timezone.localdate().isoformat()),
            ('ordering', 'folio'),
        ]
        if columnas:
            params.append(('columnas', columnas))
        return f'?{urlencode(params)}'

    def test_lista_rol1_aplica_la_seleccion_combinada(self):
        response = self.client.get(f'/api/bitacora/{self._query()}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['total'], 1)
        self.assertEqual(response.data['results'][0]['usuario'], self.responsable.email)
        self.assertEqual(response.data['results'][0]['estadoNuevo'], 'Concluido')

    def test_excel_rol1_conserva_filtros_orden_y_columnas(self):
        import io
        import openpyxl

        columnas = 'folio,nombre,usuario,estatus,estado_ant,estado_nuevo,observaciones'
        response = self.client.get(
            f'/api/bitacora/exportar/excel/{self._query(columnas)}'
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        libro = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
        filas = list(libro.active.iter_rows(values_only=True))
        encabezado = ('Folio', 'Nombre', 'Usuario', 'Estatus', 'Estado anterior', 'Estado nuevo', 'Observaciones')
        indice = next(i for i, fila in enumerate(filas) if tuple(fila[:7]) == encabezado)
        datos = [fila[:7] for fila in filas[indice + 1:] if fila[0]]
        self.assertEqual(datos, [(
            'ANAM/PARTICULAR/FUS/ROL1/001',
            'Responsable correcto',
            self.responsable.email,
            'Concluido',
            'Pendiente de validación',
            'Concluido',
            'Movimiento que debe exportarse',
        )])

    def test_pdf_rol1_conserva_la_misma_seleccion(self):
        columnas = 'folio,nombre,estatus,observaciones'
        response = self.client.get(
            f'/api/bitacora/exportar/pdf/{self._query(columnas)}'
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))


class BitacoraEquipoParticularExportTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado',
            defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.unidad = UnidadAdministrativa.objects.create(
            idUnidadAdministrativa=99002,
            clave='QA-EQUIPO',
            unidadAdministrativa='Unidad visible del equipo',
            esUnidadAdministrativa=1,
            activo=1,
        )
        cls.propietario = User.objects.create_user(
            username='dueno-equipo-bitacora@anam.gob.mx',
            email='dueno-equipo-bitacora@anam.gob.mx',
            password='x',
        )
        cls.propietario_ajeno = User.objects.create_user(
            username='dueno-ajeno-bitacora@anam.gob.mx',
            email='dueno-ajeno-bitacora@anam.gob.mx',
            password='x',
        )
        cls.equipo = User.objects.create_user(
            username='equipo-bitacora@anam.gob.mx',
            email='equipo-bitacora@anam.gob.mx',
            password='x',
        )
        cls.responsable = User.objects.create_user(
            username='responsable-equipo@anam.gob.mx',
            email='responsable-equipo@anam.gob.mx',
            password='x',
        )
        cls.responsable_ajeno = User.objects.create_user(
            username='responsable-ajeno-equipo@anam.gob.mx',
            email='responsable-ajeno-equipo@anam.gob.mx',
            password='x',
        )
        CorreoAutorizado.objects.create(
            email=cls.propietario.email, nombre='Titular particular propio', rol='ROL1', activo=1,
        )
        CorreoAutorizado.objects.create(
            email=cls.propietario_ajeno.email, nombre='Titular particular ajeno', rol='ROL1', activo=1,
        )
        CorreoAutorizado.objects.create(
            email=cls.equipo.email,
            nombre='Equipo del titular propio',
            rol='EQUIPO_PARTICULAR',
            activo=1,
            idUsuarioRegistra=cls.propietario.id,
        )
        CorreoAutorizado.objects.create(
            email=cls.responsable.email,
            nombre='Responsable visible',
            rol='ROL2',
            unidadAdministrativa=cls.unidad,
            activo=1,
        )
        CorreoAutorizado.objects.create(
            email=cls.responsable_ajeno.email,
            nombre='Responsable oculto',
            rol='ROL2',
            activo=1,
        )
        cls.fus_propio = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/EQUIPO/001',
            idSolicitanteInterno=cls.propietario,
            descripcion='FUS del titular asociado',
            contexto='',
            estatusParticular_id='Registrado',
        )
        cls.fus_ajeno = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/EQUIPO/002',
            idSolicitanteInterno=cls.propietario_ajeno,
            descripcion='FUS de otra rama',
            contexto='',
            estatusParticular_id='Registrado',
        )
        Bitacora.objects.create(
            usuario=cls.responsable.email,
            rol='ROL2',
            accion='CONCLUSION_FUS',
            fusFolio=cls.fus_propio.folio,
            estadoAnterior='Pendiente_validacion',
            estadoNuevo='Concluido',
            observaciones='Movimiento visible para el equipo',
        )
        Bitacora.objects.create(
            usuario=cls.responsable_ajeno.email,
            rol='ROL2',
            accion='CONCLUSION_FUS',
            fusFolio=cls.fus_ajeno.folio,
            estadoAnterior='Pendiente_validacion',
            estadoNuevo='Concluido',
            observaciones='Movimiento de otra rama',
        )

    def setUp(self):
        self.client.force_authenticate(self.equipo)

    def _query(self, columnas=None):
        from urllib.parse import urlencode
        from django.utils import timezone

        params = [
            ('q', 'Responsable visible'),
            ('usuario', self.responsable.email),
            ('unidadAdministrativa', str(self.unidad.idUnidadAdministrativa)),
            ('estatus_fus', 'Concluido'),
            ('fecha_desde', timezone.localdate().isoformat()),
            ('fecha_hasta', timezone.localdate().isoformat()),
            ('ordering', 'folio'),
        ]
        if columnas:
            params.append(('columnas', columnas))
        return f'?{urlencode(params)}'

    def test_equipo_solo_ve_la_rama_de_su_titular_y_aplica_filtros(self):
        response = self.client.get(f'/api/bitacora/{self._query()}')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['total'], 1)
        self.assertEqual(response.data['results'][0]['fusFolio'], self.fus_propio.folio)

    def test_opciones_de_responsable_no_exponen_otras_ramas(self):
        response = self.client.get('/api/bitacora/responsables/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual([item['email'] for item in response.data], [self.responsable.email])
        self.assertEqual(response.data[0]['unidadAdministrativa'], self.unidad.idUnidadAdministrativa)

    def test_excel_y_pdf_del_equipo_conservan_seleccion(self):
        import io
        import openpyxl

        columnas = 'folio,nombre,usuario,unidadAdministrativa,estatus,observaciones'
        excel = self.client.get(f'/api/bitacora/exportar/excel/{self._query(columnas)}')
        self.assertEqual(excel.status_code, status.HTTP_200_OK)
        libro = openpyxl.load_workbook(io.BytesIO(excel.content), read_only=True)
        filas = list(libro.active.iter_rows(values_only=True))
        encabezado = ('Folio', 'Nombre', 'Usuario', 'Unidad administrativa', 'Estatus', 'Observaciones')
        indice = next(i for i, fila in enumerate(filas) if tuple(fila[:6]) == encabezado)
        datos = [fila[:6] for fila in filas[indice + 1:] if fila[0]]
        self.assertEqual(datos, [(
            self.fus_propio.folio,
            'Responsable visible',
            self.responsable.email,
            self.unidad.unidadAdministrativa,
            'Concluido',
            'Movimiento visible para el equipo',
        )])

        pdf = self.client.get(f'/api/bitacora/exportar/pdf/{self._query(columnas)}')
        self.assertEqual(pdf.status_code, status.HTTP_200_OK)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))


class ReportesTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado',
            defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.rol1 = User.objects.create_user(
            username='reportes@anam.gob.mx', email='reportes@anam.gob.mx', password='x',
        )
        CorreoAutorizado.objects.create(
            email=cls.rol1.email, nombre='Usuario Reportes', rol='ROL1', activo=1,
        )
        cls.rol2 = User.objects.create_user(
            username='sinreportes@anam.gob.mx', email='sinreportes@anam.gob.mx', password='x',
        )
        CorreoAutorizado.objects.create(
            email=cls.rol2.email, nombre='Sin Reportes', rol='ROL2', activo=1,
        )
        FUS.objects.create(
            folio='FUS/REPORTES/001', idSolicitanteInterno=cls.rol1,
            descripcion='Solicitud para indicadores', contexto='',
            prioridad='Alta', estatusParticular_id='Registrado',
        )

    def test_rol1_consulta_resumen(self):
        self.client.force_authenticate(self.rol1)
        response = self.client.get('/api/reportes/resumen/?fecha_inicio=2020-01-01&fecha_fin=2030-01-01')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['resumen']['total'], 1)
        self.assertIn('evolucion', response.data)
        self.assertIn('detalle', response.data)

    def test_rol2_no_puede_consultar_reportes(self):
        self.client.force_authenticate(self.rol2)
        response = self.client.get('/api/reportes/resumen/')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_exportaciones_generan_archivo(self):
        self.client.force_authenticate(self.rol1)
        for formato, content_type in (
            ('pdf', 'application/pdf'),
            ('excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('pptx', 'application/vnd.openxmlformats-officedocument.presentationml.presentation'),
        ):
            response = self.client.post(
                f'/api/reportes/exportar/{formato}/?fecha_inicio=2020-01-01&fecha_fin=2030-01-01',
                {'secciones': ['resumen', 'detalle']},
                format='json',
            )
            self.assertEqual(response.status_code, status.HTTP_200_OK)
            self.assertEqual(response['Content-Type'], content_type)
            self.assertGreater(len(response.content), 100)


class FUSIDORTests(APITestCase):
    """Un usuario ROL1 no debe poder ver/editar/turnar un FUS que no le pertenece
    (IDOR): debe recibir 404, no 403 ni 200, para no confirmar la existencia del recurso."""

    @classmethod
    def setUpTestData(cls):
        cls.estatus_registrado, _ = Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.medio = MedioRecepcion.objects.create(nombreMedio='Correo electrónico', paraTurnado=1)

        cls.user_a = User.objects.create_user(username='a@anam.gob.mx', email='a@anam.gob.mx', password='x')
        CorreoAutorizado.objects.create(email='a@anam.gob.mx', nombre='Usuario A', rol='ROL1', activo=1)

        cls.user_b = User.objects.create_user(username='b@anam.gob.mx', email='b@anam.gob.mx', password='x')
        CorreoAutorizado.objects.create(email='b@anam.gob.mx', nombre='Usuario B', rol='ROL1', activo=1)

        cls.user_dest = User.objects.create_user(username='dest@anam.gob.mx', email='dest@anam.gob.mx', password='x')
        CorreoAutorizado.objects.create(email='dest@anam.gob.mx', nombre='Destinatario', rol='ROL2', activo=1)

        cls.fus_de_b = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0001/2026',
            idSolicitanteInterno=cls.user_b,
            descripcion='Solicitud de B',
            contexto='',
            estatusParticular_id='Registrado',
            idUsuarioRegistra=cls.user_b.id,
        )

    def test_get_fus_ajeno_devuelve_404(self):
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get(f'/api/fus/{self.fus_de_b.pk}/')
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_patch_fus_ajeno_devuelve_404(self):
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.patch(f'/api/fus/{self.fus_de_b.pk}/', {'descripcion': 'hackeado'})
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)
        self.fus_de_b.refresh_from_db()
        self.assertEqual(self.fus_de_b.descripcion, 'Solicitud de B')

    def test_turnar_fus_ajeno_devuelve_404(self):
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.post(f'/api/fus/{self.fus_de_b.pk}/turnar/', {
            'destinatarios': [{'idDestinatario': self.user_dest.id, 'idMedio': self.medio.id}],
            'solicitudTexto': 'intento de turnado ajeno',
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_actividad_fus_ajeno_devuelve_404(self):
        self.client.force_authenticate(user=self.user_a)
        resp = self.client.get(f'/api/fus/{self.fus_de_b.pk}/actividad/')
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_dueno_si_puede_ver_su_propio_fus(self):
        self.client.force_authenticate(user=self.user_b)
        resp = self.client.get(f'/api/fus/{self.fus_de_b.pk}/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)


class TurnarMultiplesDestinatariosTests(APITestCase):
    """TurnarFUSView — Bloque N: turnar a varios destinatarios del mismo FUS
    con el mismo medio no debe repetir la consulta a MedioRecepcion ni el
    get_or_create de la Actividad de "vence FUS" una vez por destinatario."""

    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        Estatus.objects.get_or_create(
            clave='Recibido', defaults={'nombre': 'Recibido', 'tipoFlujo': 'TITULAR', 'orden': 1},
        )
        cls.medio = MedioRecepcion.objects.create(nombreMedio='Correo electrónico', paraTurnado=1)
        cls.rol1 = User.objects.create_user(username='rol1@t.mx', email='rol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol1@t.mx', nombre='Rol1', rol='ROL1', activo=1)
        cls.dest1 = User.objects.create_user(username='dest1@t.mx', email='dest1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='dest1@t.mx', nombre='Dest1', rol='ROL2', activo=1)
        cls.dest2 = User.objects.create_user(username='dest2@t.mx', email='dest2@t.mx', password='x')
        CorreoAutorizado.objects.create(email='dest2@t.mx', nombre='Dest2', rol='ROL2', activo=1)

        cls.fus = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0400/2026', idSolicitanteInterno=cls.rol1,
            descripcion='x', contexto='', estatusParticular_id='Registrado',
            idUsuarioRegistra=cls.rol1.id,
            fechaLimite=timezone.now() + timedelta(days=3),
        )

    def test_turnar_a_dos_destinatarios_mismo_medio(self):
        self.client.force_authenticate(user=self.rol1)
        resp = self.client.post(f'/api/fus/{self.fus.pk}/turnar/', {
            'destinatarios': [
                {'idDestinatario': self.dest1.id, 'idMedio': self.medio.id},
                {'idDestinatario': self.dest2.id, 'idMedio': self.medio.id},
            ],
            'solicitudTexto': 'turnado de prueba',
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)

        turnados = Turnado.objects.filter(idFus=self.fus, activo=1)
        self.assertEqual(turnados.count(), 2)
        self.assertEqual(set(turnados.values_list('idDestinatario_id', flat=True)), {self.dest1.id, self.dest2.id})

        # Una sola Actividad "vence FUS" para el FUS, con ambos destinatarios
        # como participantes — no una por destinatario.
        actividades = Actividad.objects.filter(idFusRelacionado=self.fus, tipo='limite')
        self.assertEqual(actividades.count(), 1)
        self.assertEqual(
            set(actividades.first().participantes.values_list('id', flat=True)),
            {self.dest1.id, self.dest2.id},
        )

    def test_medio_recepcion_se_consulta_una_sola_vez_para_ambos_destinatarios(self):
        from django.test.utils import CaptureQueriesContext
        from django.db import connection

        self.client.force_authenticate(user=self.rol1)
        with CaptureQueriesContext(connection) as ctx:
            resp = self.client.post(f'/api/fus/{self.fus.pk}/turnar/', {
                'destinatarios': [
                    {'idDestinatario': self.dest1.id, 'idMedio': self.medio.id},
                    {'idDestinatario': self.dest2.id, 'idMedio': self.medio.id},
                ],
                'solicitudTexto': 'turnado de prueba',
            }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)

        consultas_medio = [q for q in ctx.captured_queries if 'scs_cat_medios_recepcion' in q['sql']]
        # SELECT del medio (una sola vez, cacheado por id) — no debe haber
        # una segunda SELECT a MedioRecepcion para el segundo destinatario.
        selects_medio = [q for q in consultas_medio if q['sql'].strip().upper().startswith('SELECT')]
        self.assertEqual(len(selects_medio), 1, consultas_medio)


class ComisionarFUSActividadTests(APITestCase):
    """ComisionarFUSView — el comisionado asignado debe quedar como
    participante de la Actividad "vence FUS" del calendario, igual que ya
    pasa con el destinatario de un Turnado (TurnarMultiplesDestinatariosTests)
    — sin esto, la temporalidad del FUS nunca aparecía en el calendario del
    comisionado, solo en el de quien lo registró/turnó."""

    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.rol1 = User.objects.create_user(username='rol1@t.mx', email='rol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol1@t.mx', nombre='Rol1', rol='ROL1', activo=1)
        cls.comisionado = User.objects.create_user(username='comi@t.mx', email='comi@t.mx', password='x')
        CorreoAutorizado.objects.create(email='comi@t.mx', nombre='Comi', rol='COMISIONADO', activo=1)

        cls.fus = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0500/2026', idSolicitanteInterno=cls.rol1,
            descripcion='x', contexto='', estatusParticular_id='Registrado',
            idUsuarioRegistra=cls.rol1.id,
            fechaLimite=timezone.now() + timedelta(days=3),
        )

    def test_comisionar_agrega_al_comisionado_como_participante(self):
        self.client.force_authenticate(user=self.rol1)
        resp = self.client.post(f'/api/fus/{self.fus.pk}/comisionar/', {
            'comisionado_id': self.comisionado.id,
        }, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK, resp.data)

        actividad = Actividad.objects.get(idFusRelacionado=self.fus, tipo='limite', activo=1)
        self.assertIn(self.comisionado.id, actividad.participantes.values_list('id', flat=True))


class _FixtureRolesFUS(APITestCase):
    """Base con un FUS y un usuario por cada combinación autorizado/ajeno de
    los 4 roles que pueden llegar a DescargarEvidenciaView/FUSTrazabilidadView
    — reusada por ambas suites para no duplicar el armado de datos."""

    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.medio = MedioRecepcion.objects.create(nombreMedio='Correo electrónico', paraTurnado=1)

        cls.rol1 = User.objects.create_user(username='rol1@t.mx', email='rol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol1@t.mx', nombre='Rol1', rol='ROL1', activo=1)

        cls.otro_rol1 = User.objects.create_user(username='otrorol1@t.mx', email='otrorol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='otrorol1@t.mx', nombre='OtroRol1', rol='ROL1', activo=1)

        # EQUIPO_PARTICULAR autorizado: idUsuarioRegistra apunta al dueño real del FUS.
        cls.equipo_de_rol1 = User.objects.create_user(username='equipo@t.mx', email='equipo@t.mx', password='x')
        CorreoAutorizado.objects.create(
            email='equipo@t.mx', nombre='Equipo', rol='EQUIPO_PARTICULAR', activo=1,
            idUsuarioRegistra=cls.rol1.id,
        )
        # EQUIPO_PARTICULAR ajeno: asiste a otro_rol1, no al dueño del FUS de prueba.
        cls.equipo_ajeno = User.objects.create_user(username='equipoajeno@t.mx', email='equipoajeno@t.mx', password='x')
        CorreoAutorizado.objects.create(
            email='equipoajeno@t.mx', nombre='EquipoAjeno', rol='EQUIPO_PARTICULAR', activo=1,
            idUsuarioRegistra=cls.otro_rol1.id,
        )

        cls.rol2_destinatario = User.objects.create_user(username='rol2dest@t.mx', email='rol2dest@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol2dest@t.mx', nombre='Rol2Dest', rol='ROL2', activo=1)

        cls.rol2_ajeno = User.objects.create_user(username='rol2ajeno@t.mx', email='rol2ajeno@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol2ajeno@t.mx', nombre='Rol2Ajeno', rol='ROL2', activo=1)

        cls.comisionado_asignado = User.objects.create_user(username='comi@t.mx', email='comi@t.mx', password='x')
        CorreoAutorizado.objects.create(email='comi@t.mx', nombre='Comi', rol='COMISIONADO', activo=1)

        cls.comisionado_ajeno = User.objects.create_user(username='comiajeno@t.mx', email='comiajeno@t.mx', password='x')
        CorreoAutorizado.objects.create(email='comiajeno@t.mx', nombre='ComiAjeno', rol='COMISIONADO', activo=1)

        cls.fus = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0100/2026',
            idSolicitanteInterno=cls.rol1,
            descripcion='Solicitud de prueba', contexto='',
            estatusParticular_id='Registrado',
            idUsuarioRegistra=cls.rol1.id,
            idComisionado=cls.comisionado_asignado,
        )
        Turnado.objects.create(
            idFus=cls.fus, idRemitente=cls.rol1, idDestinatario=cls.rol2_destinatario,
            idMedio=cls.medio, estatusTitular_id='Recibido', activo=1,
        )


class DescargarEvidenciaViewTests(_FixtureRolesFUS):
    """ROL1 ve cualquier evidencia; EQUIPO_PARTICULAR solo la de su ROL1
    asociado; ROL2 solo si tiene un Turnado activo dirigido a él; COMISIONADO
    solo si el FUS le está asignado. Cualquier otro caso debe ser 404 (nunca
    403, para no confirmar que el recurso existe) — y la ruta resuelta del
    archivo nunca debe poder salirse de MEDIA_ROOT."""

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.tmp_media = tempfile.mkdtemp()
        ruta_rel = os.path.join('evidencias', 'test', 'archivo.txt')
        ruta_abs = os.path.join(cls.tmp_media, ruta_rel)
        os.makedirs(os.path.dirname(ruta_abs), exist_ok=True)
        with open(ruta_abs, 'wb') as f:
            f.write(b'contenido de prueba')

        cls.evidencia = Evidencia.objects.create(
            idFus=cls.fus, nombreArchivo='archivo.txt', rutaArchivo=ruta_rel,
            tipoMime='text/plain', idUsuarioRegistra=cls.rol1.id,
        )

    def _get(self, user, evidencia=None):
        evidencia = evidencia or self.evidencia
        with override_settings(MEDIA_ROOT=self.tmp_media):
            self.client.force_authenticate(user=user)
            return self.client.get(f'/api/evidencias/{evidencia.id}/descargar/')

    def test_rol1_ve_cualquier_evidencia(self):
        self.assertEqual(self._get(self.rol1).status_code, status.HTTP_200_OK)

    def test_equipo_particular_de_ese_rol1_autorizado(self):
        self.assertEqual(self._get(self.equipo_de_rol1).status_code, status.HTTP_200_OK)

    def test_equipo_particular_ajeno_404(self):
        self.assertEqual(self._get(self.equipo_ajeno).status_code, status.HTTP_404_NOT_FOUND)

    def test_rol2_destinatario_autorizado(self):
        self.assertEqual(self._get(self.rol2_destinatario).status_code, status.HTTP_200_OK)

    def test_rol2_ajeno_404(self):
        self.assertEqual(self._get(self.rol2_ajeno).status_code, status.HTTP_404_NOT_FOUND)

    def test_comisionado_asignado_autorizado(self):
        self.assertEqual(self._get(self.comisionado_asignado).status_code, status.HTTP_200_OK)

    def test_comisionado_no_asignado_404(self):
        self.assertEqual(self._get(self.comisionado_ajeno).status_code, status.HTTP_404_NOT_FOUND)

    def test_ruta_fuera_de_media_root_devuelve_404(self):
        # rutaArchivo corrupta/manipulada que intenta escapar de MEDIA_ROOT —
        # aunque el usuario esté autorizado a ver el FUS, el archivo no se sirve.
        evidencia_maliciosa = Evidencia.objects.create(
            idFus=self.fus, nombreArchivo='fuera.txt',
            rutaArchivo=os.path.join('..', '..', 'etc', 'passwd'),
            tipoMime='text/plain', idUsuarioRegistra=self.rol1.id,
        )
        resp = self._get(self.rol1, evidencia=evidencia_maliciosa)
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)


class FUSTrazabilidadViewTests(_FixtureRolesFUS):
    """Mismas reglas de acceso que DescargarEvidenciaView: un usuario sin
    relación con el FUS debe recibir 404 antes de que se calcule/devuelva
    cualquier evento de la línea de tiempo."""

    def _get(self, user):
        self.client.force_authenticate(user=user)
        return self.client.get(f'/api/fus/trazabilidad/{self.fus.folio}/')

    def test_rol1_ve_cualquier_fus(self):
        self.assertEqual(self._get(self.rol1).status_code, status.HTTP_200_OK)

    def test_equipo_particular_de_ese_rol1_autorizado(self):
        self.assertEqual(self._get(self.equipo_de_rol1).status_code, status.HTTP_200_OK)

    def test_equipo_particular_ajeno_404(self):
        self.assertEqual(self._get(self.equipo_ajeno).status_code, status.HTTP_404_NOT_FOUND)

    def test_rol2_destinatario_autorizado(self):
        self.assertEqual(self._get(self.rol2_destinatario).status_code, status.HTTP_200_OK)

    def test_rol2_ajeno_404(self):
        self.assertEqual(self._get(self.rol2_ajeno).status_code, status.HTTP_404_NOT_FOUND)

    def test_comisionado_asignado_autorizado(self):
        self.assertEqual(self._get(self.comisionado_asignado).status_code, status.HTTP_200_OK)

    def test_comisionado_no_asignado_404(self):
        self.assertEqual(self._get(self.comisionado_ajeno).status_code, status.HTTP_404_NOT_FOUND)


class FUSDetalleAuditoriaViewTests(_FixtureRolesFUS):
    """El modal de detalle respeta la relación del usuario con el FUS."""

    def _get(self, user):
        self.client.force_authenticate(user=user)
        return self.client.get(f'/api/fus/detalle-auditoria/{self.fus.folio}/')

    def test_rol2_destinatario_puede_ver_detalle(self):
        response = self._get(self.rol2_destinatario)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['folio'], self.fus.folio)

    def test_rol2_ajeno_no_puede_ver_detalle(self):
        self.assertEqual(
            self._get(self.rol2_ajeno).status_code,
            status.HTTP_404_NOT_FOUND,
        )


class ValidacionPorPersonaTurnadoTests(APITestCase):
    """FUS turnado a varias personas (sin comisionado): cada Titular avanza
    su propio turnado (responder -> marcar atendido) de forma independiente,
    y el Particular rechaza/concluye la parte de UNA persona sin afectar a
    las demás — el FUS solo pasa a 'Concluido' cuando TODOS los turnados
    activos ya están concluidos."""

    @classmethod
    def setUpTestData(cls):
        for clave in ('Turnado', 'En_seguimiento', 'Atendido', 'Pendiente_validacion', 'Concluido', 'Rechazado'):
            Estatus.objects.get_or_create(
                clave=clave, defaults={'nombre': clave, 'tipoFlujo': 'PARTICULAR', 'orden': 1},
            )
        cls.medio = MedioRecepcion.objects.create(nombreMedio='Correo electrónico', paraTurnado=1)

        cls.rol1 = User.objects.create_user(username='rol1@t.mx', email='rol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol1@t.mx', nombre='Rol1', rol='ROL1', activo=1)

        cls.mariana = User.objects.create_user(username='mariana@t.mx', email='mariana@t.mx', password='x')
        CorreoAutorizado.objects.create(email='mariana@t.mx', nombre='Mariana', rol='ROL2', activo=1)

        cls.lucia = User.objects.create_user(username='lucia@t.mx', email='lucia@t.mx', password='x')
        CorreoAutorizado.objects.create(email='lucia@t.mx', nombre='Lucía', rol='ROL2', activo=1)

        cls.fus = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0200/2026',
            idSolicitanteInterno=cls.rol1,
            descripcion='Solicitud turnada a dos personas', contexto='',
            estatusParticular_id='Turnado',
            idUsuarioRegistra=cls.rol1.id,
        )
        cls.t_mariana = Turnado.objects.create(
            idFus=cls.fus, idRemitente=cls.rol1, idDestinatario=cls.mariana,
            idMedio=cls.medio, estatusTitular_id='Recibido', activo=1,
        )
        cls.t_lucia = Turnado.objects.create(
            idFus=cls.fus, idRemitente=cls.rol1, idDestinatario=cls.lucia,
            idMedio=cls.medio, estatusTitular_id='Recibido', activo=1,
        )

    def _responder(self, user, turnado):
        self.client.force_authenticate(user=user)
        return self.client.post(
            f'/api/turnados/{turnado.id}/seguimientos/',
            {'descripcionActividad': 'Reviso el caso', 'accionTexto': ''},
        )

    def _atendido(self, user, turnado):
        self.client.force_authenticate(user=user)
        return self.client.post(f'/api/turnados/{turnado.id}/atendido/')

    def _concluir_persona(self, user, turnado):
        self.client.force_authenticate(user=user)
        return self.client.post(f'/api/turnados/{turnado.id}/concluir-persona/')

    def _rechazar_persona(self, user, turnado, motivo='No es suficiente'):
        self.client.force_authenticate(user=user)
        return self.client.post(f'/api/turnados/{turnado.id}/rechazar-persona/', {'motivo': motivo})

    def test_atendido_solo_afecta_su_propio_turnado(self):
        self._responder(self.mariana, self.t_mariana)
        resp = self._atendido(self.mariana, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        self.t_mariana.refresh_from_db()
        self.t_lucia.refresh_from_db()
        self.fus.refresh_from_db()
        self.assertEqual(self.t_mariana.estatusTitular_id, 'Pendiente_validacion')
        self.assertEqual(self.t_lucia.estatusTitular_id, 'Recibido')  # sin tocar
        self.assertEqual(self.fus.estatusParticular_id, 'Atendido')

    def test_rol2_no_puede_marcar_atendido_sin_responder_antes(self):
        resp = self._atendido(self.mariana, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_rol2_ajeno_no_puede_marcar_atendido_de_otro_turnado(self):
        self._responder(self.mariana, self.t_mariana)
        resp = self._atendido(self.lucia, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)

    def test_rol1_no_puede_concluir_antes_de_atendido(self):
        self._responder(self.mariana, self.t_mariana)
        resp = self._concluir_persona(self.rol1, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_rol2_no_puede_concluir_la_parte_de_otra_persona(self):
        self._responder(self.mariana, self.t_mariana)
        self._atendido(self.mariana, self.t_mariana)
        resp = self._concluir_persona(self.mariana, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)

    def test_fus_sigue_atendido_hasta_que_todas_las_personas_concluyen(self):
        # Mariana responde, marca atendido, Rol 1 la concluye — Lucía sigue pendiente.
        self._responder(self.mariana, self.t_mariana)
        self._atendido(self.mariana, self.t_mariana)
        resp = self._concluir_persona(self.rol1, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        self.t_mariana.refresh_from_db()
        self.fus.refresh_from_db()
        self.assertEqual(self.t_mariana.estatusTitular_id, 'Concluido')
        self.assertEqual(self.fus.estatusParticular_id, 'Atendido')  # todavía no, falta Lucía

        # Lucía responde, marca atendido, Rol 1 la concluye -> ahora sí el FUS completo
        self._responder(self.lucia, self.t_lucia)
        self._atendido(self.lucia, self.t_lucia)
        resp = self._concluir_persona(self.rol1, self.t_lucia)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        self.t_lucia.refresh_from_db()
        self.fus.refresh_from_db()
        self.assertEqual(self.t_lucia.estatusTitular_id, 'Concluido')
        self.assertEqual(self.fus.estatusParticular_id, 'Concluido')
        self.assertIsNotNone(self.fus.fechaConclusion)

    def test_rechazar_persona_solo_afecta_su_propio_turnado(self):
        self._responder(self.mariana, self.t_mariana)
        self._atendido(self.mariana, self.t_mariana)
        resp = self._rechazar_persona(self.rol1, self.t_mariana, motivo='Falta información')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

        self.t_mariana.refresh_from_db()
        self.t_lucia.refresh_from_db()
        self.fus.refresh_from_db()
        self.assertEqual(self.t_mariana.estatusTitular_id, 'Rechazado')
        self.assertEqual(self.t_lucia.estatusTitular_id, 'Recibido')       # ajena, sin tocar
        self.assertEqual(self.fus.estatusParticular_id, 'Atendido')        # el FUS no se rechaza completo

    def test_rechazar_persona_requiere_motivo(self):
        self._responder(self.mariana, self.t_mariana)
        self._atendido(self.mariana, self.t_mariana)
        self.client.force_authenticate(user=self.rol1)
        resp = self.client.post(f'/api/turnados/{self.t_mariana.id}/rechazar-persona/', {'motivo': '  '})
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_rol2_ajeno_no_puede_validar(self):
        self._responder(self.mariana, self.t_mariana)
        self._atendido(self.mariana, self.t_mariana)
        resp = self._concluir_persona(self.lucia, self.t_mariana)
        self.assertEqual(resp.status_code, status.HTTP_403_FORBIDDEN)


class BusquedaFUSFulltextTests(APITransactionTestCase):
    """FUSListCreateView.get(search=...) — migración 0033: MATCH()...AGAINST()
    IN BOOLEAN MODE sobre FUS.descripcion/contexto, Evidencia.nombreArchivo/
    comentarios y Turnado.solicitudTexto, con respaldo a icontains para
    términos más cortos que innodb_ft_min_token_size.

    APITransactionTestCase (no APITestCase): InnoDB solo refleja filas
    nuevas en el índice FULLTEXT después de un COMMIT real — con
    APITestCase (que envuelve cada test en una transacción que hace
    rollback al final) MATCH() nunca ve los datos de setUp, aunque
    icontains sí los vería vía MVCC dentro de esa misma transacción."""

    def setUp(self):
        Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        Estatus.objects.get_or_create(
            clave='Turnado', defaults={'nombre': 'Turnado', 'tipoFlujo': 'PARTICULAR', 'orden': 2},
        )
        Estatus.objects.get_or_create(
            clave='Recibido', defaults={'nombre': 'Recibido', 'tipoFlujo': 'TITULAR', 'orden': 1},
        )
        self.medio = MedioRecepcion.objects.create(nombreMedio='Correo electrónico', paraTurnado=1)
        self.rol1 = User.objects.create_user(username='rol1@t.mx', email='rol1@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol1@t.mx', nombre='Rol1', rol='ROL1', activo=1)
        self.rol2 = User.objects.create_user(username='rol2@t.mx', email='rol2@t.mx', password='x')
        CorreoAutorizado.objects.create(email='rol2@t.mx', nombre='Rol2', rol='ROL2', activo=1)

        self.fus_descripcion = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0200/2026', idSolicitanteInterno=self.rol1,
            descripcion='Solicitud sobre credenciales de acceso institucional', contexto='',
            estatusParticular_id='Registrado', idUsuarioRegistra=self.rol1.id,
        )
        self.fus_evidencia = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0201/2026', idSolicitanteInterno=self.rol1,
            descripcion='Otro asunto sin relación', contexto='',
            estatusParticular_id='Registrado', idUsuarioRegistra=self.rol1.id,
        )
        Evidencia.objects.create(
            idFus=self.fus_evidencia, nombreArchivo='comprobante_pago.pdf',
            comentarios='Incluye el recibo de pago correspondiente', activo=1,
        )
        self.fus_turnado = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0202/2026', idSolicitanteInterno=self.rol1,
            descripcion='Asunto distinto', contexto='',
            estatusParticular_id='Turnado', idUsuarioRegistra=self.rol1.id,
        )
        Turnado.objects.create(
            idFus=self.fus_turnado, idRemitente=self.rol1, idDestinatario=self.rol2,
            idMedio=self.medio, estatusTitular_id='Recibido', activo=1,
            solicitudTexto='Favor de atender la revisión aduanera solicitada',
        )

    def _buscar(self, termino):
        self.client.force_authenticate(user=self.rol1)
        resp = self.client.get('/api/fus/', {'search': termino})
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        return {f['folio'] for f in resp.data['results']}

    def test_encuentra_por_descripcion(self):
        self.assertIn(self.fus_descripcion.folio, self._buscar('credenciales'))

    def test_encuentra_por_evidencia(self):
        self.assertIn(self.fus_evidencia.folio, self._buscar('comprobante'))

    def test_encuentra_por_turnado(self):
        self.assertIn(self.fus_turnado.folio, self._buscar('aduanera'))

    def test_termino_corto_usa_respaldo_icontains(self):
        # 'de' tiene 2 caracteres — más corto que innodb_ft_min_token_size,
        # así que no debe usar MATCH() (no encontraría nada) sino icontains.
        folios = self._buscar('de')
        self.assertIn(self.fus_descripcion.folio, folios)

    def test_termino_sin_coincidencias_no_falla(self):
        self.assertEqual(self._buscar('xyznoexisteenningunlado'), set())

    def test_caracteres_de_operador_booleano_no_rompen_la_busqueda(self):
        # +, -, *, ", (, ) tienen significado especial en modo booleano de
        # MySQL — deben limpiarse antes de armar el término, no llegar
        # crudos a MATCH()...AGAINST() (que fallaría con sintaxis inválida).
        folios = self._buscar('+credenciales* -"acceso"')
        self.assertIn(self.fus_descripcion.folio, folios)


class NotificarPorCorreoTests(APITestCase):
    """notificar_por_correo() — Bloque L: con RQ_QUEUES configurado (Redis
    disponible) encola el envío con reintento automático (django-rq); sin
    Redis cae al hilo suelto de respaldo — mismo criterio condicional que
    CHANNEL_LAYERS/CACHES en settings.py. No hay Redis real en este entorno
    de pruebas, así que la ruta con cola se verifica mockeando
    django_rq.get_queue en vez de correr un worker de verdad."""

    @classmethod
    def setUpTestData(cls):
        Estatus.objects.get_or_create(
            clave='Registrado', defaults={'nombre': 'Registrado', 'tipoFlujo': 'PARTICULAR', 'orden': 1},
        )
        cls.user = User.objects.create_user(username='dest@t.mx', email='dest@t.mx', password='x')
        cls.otro_user = User.objects.create_user(username='dest2@t.mx', email='dest2@t.mx', password='x')

    def _crear_notificacion(self):
        return Notificacion.objects.create(
            idDestinatario=self.user, fusFolio='', tipoEvento='ACTIVIDAD', mensaje='prueba',
        )

    def test_sin_rq_usa_hilo_de_respaldo(self):
        from solicitudes.services.notificaciones import notificar_por_correo
        notif = self._crear_notificacion()
        with patch('solicitudes.services.notificaciones.threading.Thread') as MockThread:
            with self.captureOnCommitCallbacks(execute=True):
                notificar_por_correo(notif)
            MockThread.assert_called_once()
            _, kwargs = MockThread.call_args
            # Se pasan ids en una lista (notificar_por_correo_lote), no el
            # objeto ni un id suelto — ver _enviar_correos_lote().
            self.assertEqual(kwargs['args'], ([notif.id],))

    @override_settings(RQ_QUEUES={'default': {'URL': 'redis://localhost:6379/0'}})
    def test_con_rq_encola_con_reintento(self):
        from solicitudes.services.notificaciones import notificar_por_correo
        notif = self._crear_notificacion()
        mock_queue = MagicMock()
        with patch('django_rq.get_queue', return_value=mock_queue):
            with self.captureOnCommitCallbacks(execute=True):
                notificar_por_correo(notif)
        mock_queue.enqueue.assert_called_once()
        args, kwargs = mock_queue.enqueue.call_args
        self.assertEqual(args[1], [notif.id])
        self.assertEqual(kwargs['retry'].max, 3)
        self.assertEqual(kwargs['retry'].intervals, [10, 30, 60])

    def test_lote_reutiliza_una_sola_consulta_al_fus(self):
        # Bloque M: turnar (u otra acción) a varios destinatarios del mismo
        # FUS debe resolver el FUS una sola vez para todo el lote, no una
        # vez por destinatario.
        from solicitudes.services import notificaciones as notif_mod

        fus = FUS.objects.create(
            folio='ANAM/PARTICULAR/FUS/0300/2026', idSolicitanteInterno=self.user,
            descripcion='x', contexto='', estatusParticular_id='Registrado',
            idUsuarioRegistra=self.user.id,
        )
        n1 = Notificacion.objects.create(
            idDestinatario=self.user, fusFolio=fus.folio, tipoEvento='TURNADO', mensaje='a',
        )
        n2 = Notificacion.objects.create(
            idDestinatario=self.otro_user, fusFolio=fus.folio, tipoEvento='TURNADO', mensaje='b',
        )

        # connections.close_all() también se mockea: _enviar_correos_lote lo
        # llama en su finally (correcto en un hilo/worker real de
        # producción), pero aquí se invoca en el mismo hilo/conexión del
        # test — cerrarla de verdad rompería la transacción compartida que
        # usa TestCase para los demás tests de esta clase.
        with patch.object(
            notif_mod, '_cargar_fus_para_correo', wraps=notif_mod._cargar_fus_para_correo,
        ) as mock_cargar, patch.object(notif_mod, 'EmailMultiAlternatives') as MockEmail, \
                patch.object(notif_mod.connections, 'close_all'), \
                patch('solicitudes.views.fus.generar_pdf_fus', return_value=b'%PDF-'):
            MockEmail.return_value = MagicMock()
            notif_mod._enviar_correos_lote([n1.id, n2.id])

        mock_cargar.assert_called_once_with(fus.folio)
        self.assertEqual(MockEmail.call_count, 2)
