"""Reportes — versión para el Titular/Enlace Estratégico (ROL2).

A diferencia de reportes.py (ROL1: FUS propios, carga de trabajo del equipo,
unidades administrativas), aquí el universo de datos es "mis turnados" — lo
mismo que ya consulta MisTurnadosView en turnado.py — y las secciones se
ajustan a lo que un Titular individual puede accionar: no hay "carga de
trabajo por responsable" (él ES el responsable) ni "FUS por unidad" (una sola
unidad, la suya). En cambio sí interesa prioridad (para priorizar su propia
bandeja) y tiempos de respuesta/conclusión (su propio desempeño).

Reutiliza de reportes.py todo lo que es agnóstico de "quién es el dueño de
los datos": rango de fechas, comparación contra periodo anterior, paleta de
colores/formato de Excel — para no duplicar esa lógica ya probada.
"""
import os
from collections import Counter, defaultdict
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import Turnado, Seguimiento
from ..utils import get_rol, resolver_nombre
from .reportes import (
    PALETA_HEX, VERDE_INSTITUCIONAL, LOGO_PATH,
    _etiqueta_campo, _fecha_inicio_fin, _periodo_anterior, _etiqueta_periodo,
    _delta_pct, CAMPOS_COMPARABLES, _CAMPOS_PORCENTAJE, _FORMATO_PORCENTAJE,
    _hex_color, _pdf_grafica, _pdf_membrete, _pdf_construir_numbered_canvas,
    _pdf_tabla_ejecutiva,
)

SECCIONES_TITULAR = {
    'resumen': 'Resumen ejecutivo',
    'evolucion': 'Evolución mensual de turnados',
    'estados': 'Distribución por estado',
    'prioridad': 'Distribución por prioridad',
    'tiempos': 'Distribución de tiempos de respuesta',
    'detalle': 'Detalle de turnados',
}

ESTADOS_TITULAR = ['Recibido', 'En_seguimiento', 'Pendiente_validacion', 'Concluido', 'Rechazado']


def _queryset_titular(user):
    if get_rol(user) != 'ROL2':
        raise PermissionDenied('No autorizado para consultar reportes.')
    return Turnado.objects.filter(idDestinatario=user, activo=1).select_related(
        'idFus', 'idFus__idSolicitanteInterno',
    )


def _aplicar_filtros_titular(qs, params):
    inicio, fin = _fecha_inicio_fin(params)
    qs = qs.filter(fechaRegistro__range=(inicio, fin))
    if params.get('estatus'):
        qs = qs.filter(estatusTitular_id__in=[x for x in params['estatus'].split(',') if x])
    if params.get('prioridad'):
        qs = qs.filter(idFus__prioridad__in=[x for x in params['prioridad'].split(',') if x])
    return qs.order_by('fechaRegistro'), inicio, fin


def _construir_reporte_titular(qs, inicio, fin):
    registros = list(qs)
    ahora = timezone.now()
    estados = Counter()
    meses = defaultdict(lambda: {'recibidos': 0, 'concluidos': 0, 'vencidos': 0})
    prioridades = Counter()
    tiempos = Counter({'hasta_1': 0, '1_3': 0, '3_7': 0, 'mas_7': 0})
    total_dias_conclusion = 0
    concluidos_con_tiempo = 0
    total_dias_respuesta = 0
    con_respuesta = 0
    detalle = []

    # Primera respuesta por turnado (el turnado no trae su propio timestamp
    # de "primera respuesta" — se resuelve con el Seguimiento más antiguo).
    turnado_ids = [t.id for t in registros]
    primera_respuesta = {}
    if turnado_ids:
        for s in Seguimiento.objects.filter(idTurnado_id__in=turnado_ids, activo=1).order_by('idTurnado_id', 'fechaRegistro'):
            primera_respuesta.setdefault(s.idTurnado_id, s.fechaRegistro)

    for t in registros:
        estado = t.estatusTitular_id or 'Sin estado'
        estados[estado] += 1
        fus = t.idFus
        prioridad = fus.prioridad or 'Sin prioridad'
        prioridades[prioridad] += 1

        clave_mes = timezone.localtime(t.fechaRegistro).strftime('%Y-%m') if t.fechaRegistro else 'Sin fecha'
        meses[clave_mes]['recibidos'] += 1

        vencido = bool(fus.fechaLimite and fus.fechaLimite < ahora and estado != 'Concluido')
        if vencido:
            meses[clave_mes]['vencidos'] += 1

        # Turnado no tiene columna propia de "fecha de conclusión" — se usa
        # fechaModificacion como aproximación: una vez Concluido, nada vuelve
        # a tocar el registro, así que ese timestamp SÍ es el de conclusión.
        if estado == 'Concluido' and t.fechaRegistro and t.fechaModificacion:
            meses[timezone.localtime(t.fechaModificacion).strftime('%Y-%m')]['concluidos'] += 1
            dias = max((t.fechaModificacion - t.fechaRegistro).total_seconds() / 86400, 0)
            total_dias_conclusion += dias
            concluidos_con_tiempo += 1

        primera = primera_respuesta.get(t.id)
        if primera and t.fechaRegistro:
            dias_resp = max((primera - t.fechaRegistro).total_seconds() / 86400, 0)
            total_dias_respuesta += dias_resp
            con_respuesta += 1
            if dias_resp <= 1:
                tiempos['hasta_1'] += 1
            elif dias_resp <= 3:
                tiempos['1_3'] += 1
            elif dias_resp <= 7:
                tiempos['3_7'] += 1
            else:
                tiempos['mas_7'] += 1

        detalle.append({
            'folio': fus.folio,
            'fecha_recibido': timezone.localtime(t.fechaRegistro).strftime('%d/%m/%Y') if t.fechaRegistro else '',
            'estado': estado,
            'prioridad': prioridad,
            'solicitante': resolver_nombre(fus.idSolicitanteInterno) if fus.idSolicitanteInterno_id else 'Sin solicitante',
            'fecha_limite': timezone.localtime(fus.fechaLimite).strftime('%d/%m/%Y') if fus.fechaLimite else '',
            'vencido': vencido,
        })

    total = len(registros)
    concluidos = estados.get('Concluido', 0)
    pendientes = total - concluidos
    vencidos = sum(1 for d in detalle if d['vencido'])
    dentro_sla = sum(
        1 for t in registros
        if t.estatusTitular_id == 'Concluido' and (not t.idFus.fechaLimite or t.fechaModificacion <= t.idFus.fechaLimite)
    )

    return {
        'periodo': {'inicio': inicio.date().isoformat(), 'fin': fin.date().isoformat()},
        'resumen': {
            'total': total, 'concluidos': concluidos, 'pendientes': pendientes,
            'vencidos': vencidos,
            'tiempo_promedio_respuesta': round(total_dias_respuesta / con_respuesta, 1) if con_respuesta else 0,
            'tiempo_promedio_conclusion': round(total_dias_conclusion / concluidos_con_tiempo, 1) if concluidos_con_tiempo else 0,
            'cumplimiento_sla': round(dentro_sla * 100 / concluidos, 1) if concluidos else 0,
            'tasa_conclusion': round(concluidos * 100 / total, 1) if total else 0,
        },
        'evolucion': [{'periodo': k, **v} for k, v in sorted(meses.items())],
        'estados': [{'nombre': k, 'cantidad': v, 'porcentaje': round(v * 100 / total, 1) if total else 0} for k, v in estados.most_common()],
        'prioridad': [{'nombre': p, 'cantidad': prioridades.get(p, 0)} for p in ('Alta', 'Media', 'Baja') if prioridades.get(p, 0) or total],
        'tiempos': [
            {'rango': '≤ 1 día', 'cantidad': tiempos['hasta_1']},
            {'rango': '1 - 3 días', 'cantidad': tiempos['1_3']},
            {'rango': '3 - 7 días', 'cantidad': tiempos['3_7']},
            {'rango': '> 7 días', 'cantidad': tiempos['mas_7']},
        ],
        'detalle': detalle,
    }


def _datos_titular(request):
    qs_base = _queryset_titular(request.user)
    qs, inicio, fin = _aplicar_filtros_titular(qs_base, request.query_params)
    reporte = _construir_reporte_titular(qs, inicio, fin)

    inicio_prev, fin_prev = _periodo_anterior(inicio, fin, request.query_params.get('comparar_con'))
    filtros_prev = request.query_params.dict()
    filtros_prev['fecha_inicio'] = inicio_prev.date().isoformat()
    filtros_prev['fecha_fin'] = fin_prev.date().isoformat()
    qs_prev, _, _ = _aplicar_filtros_titular(qs_base, filtros_prev)
    reporte_prev = _construir_reporte_titular(qs_prev, inicio_prev, fin_prev)

    reporte['comparacion'] = {
        'etiqueta': _etiqueta_periodo(inicio_prev, fin_prev),
        'deltas': {campo: _delta_pct(reporte['resumen'][campo], reporte_prev['resumen'][campo]) for campo in CAMPOS_COMPARABLES},
        'anterior': {campo: reporte_prev['resumen'][campo] for campo in CAMPOS_COMPARABLES},
    }
    return reporte


class ReporteTitularResumenView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(_datos_titular(request))


class ReporteTitularOpcionesView(APIView):
    """Sin unidades/responsables que elegir (el Titular es un único
    responsable de una única unidad) — solo confirma los valores válidos de
    estado, para que el frontend no los tenga que hardcodear a ciegas."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        _queryset_titular(request.user)
        return Response({
            'estados': [{'id': x, 'nombre': x.replace('_', ' ')} for x in ESTADOS_TITULAR],
            'prioridades': [{'id': x, 'nombre': x} for x in ('Alta', 'Media', 'Baja')],
        })


def _excel_grafica_titular(ws, titulo, n_filas):
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    ultima_fila = n_filas + 1
    ancla = f"A{ultima_fila + 3}"

    if titulo == 'Evolución':
        chart = LineChart()
        chart.title = 'Evolución mensual de turnados'
        chart.style = 2
        chart.height, chart.width = 9, 18
        chart.add_data(Reference(ws, min_col=2, max_col=4, min_row=1, max_row=ultima_fila), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ultima_fila))
        for serie, color in zip(chart.series, PALETA_HEX):
            serie.graphicalProperties.line.solidFill = color
            serie.graphicalProperties.line.width = 22000
        ws.add_chart(chart, ancla)
    elif titulo == 'Estados':
        from openpyxl.chart.marker import DataPoint
        from openpyxl.chart.shapes import GraphicalProperties
        chart = PieChart()
        chart.title = 'Distribución por estado'
        chart.height, chart.width = 9, 14
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ultima_fila), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ultima_fila))
        chart.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(solidFill=PALETA_HEX[i % len(PALETA_HEX)]))
            for i in range(n_filas)
        ]
        ws.add_chart(chart, ancla)
    elif titulo in ('Prioridad', 'Tiempos'):
        chart = BarChart()
        chart.title = 'Distribución por prioridad' if titulo == 'Prioridad' else 'Distribución de tiempos de respuesta'
        chart.style = 10
        chart.height, chart.width = 9, 14
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ultima_fila), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ultima_fila))
        chart.series[0].graphicalProperties.solidFill = PALETA_HEX[0 if titulo == 'Prioridad' else 1]
        ws.add_chart(chart, ancla)


def _excel_portada_titular(wb, datos, request):
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet('Portada', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 90

    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.width, img.height = 267, 45
        ws.add_image(img, 'B2')

    ws['B7'] = 'Reporte de turnados — Titular'
    ws['B7'].font = Font(bold=True, size=20, color=VERDE_INSTITUCIONAL)
    ws['B8'] = 'Inteligencia operativa — Sistema de Control de Solicitudes'
    ws['B8'].font = Font(size=11, color='666666')

    fila = 10
    nombre_titular = resolver_nombre(request.user)
    filas_info = [
        ('Titular', nombre_titular),
        ('Periodo', f"{datos['periodo']['inicio']} a {datos['periodo']['fin']}"),
    ]
    filtros_aplicados = [
        f"{etiqueta}: {request.query_params.get(clave)}"
        for clave, etiqueta in [('estatus', 'Estado'), ('prioridad', 'Prioridad')]
        if request.query_params.get(clave)
    ]
    if filtros_aplicados:
        filas_info.append(('Filtros aplicados', ' · '.join(filtros_aplicados)))
    filas_info.append(('Generado el', timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')))

    for etiqueta, valor in filas_info:
        ws.cell(row=fila, column=2, value=etiqueta).font = Font(bold=True, size=10, color=VERDE_INSTITUCIONAL)
        fila += 1
        celda = ws.cell(row=fila, column=2, value=valor)
        celda.font = Font(size=10, color='333333')
        celda.alignment = Alignment(wrap_text=True, vertical='top')
        fila += 2
    return ws


class ReporteTitularExportarExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        datos = _datos_titular(request)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        _excel_portada_titular(wb, datos, request)

        ws = wb.create_sheet('Resumen')
        ws.append(['Indicador', 'Valor'])
        for clave, valor in datos['resumen'].items():
            ws.append([_etiqueta_campo(clave), valor])
            if clave in _CAMPOS_PORCENTAJE:
                ws.cell(row=ws.max_row, column=2).number_format = _FORMATO_PORCENTAJE

        tablas = {
            'Evolución': datos['evolucion'],
            'Estados': datos['estados'],
            'Prioridad': datos['prioridad'],
            'Tiempos': datos['tiempos'],
            'Detalle': datos['detalle'],
        }
        for titulo, filas in tablas.items():
            ws = wb.create_sheet(titulo[:31])
            if filas:
                claves_col = list(filas[0])
                ws.append([_etiqueta_campo(k) for k in claves_col])
                for fila in filas:
                    ws.append(list(fila.values()))
                for idx, campo in enumerate(claves_col, start=1):
                    if campo in _CAMPOS_PORCENTAJE:
                        for celda in ws[openpyxl.utils.get_column_letter(idx)][1:]:
                            celda.number_format = _FORMATO_PORCENTAJE
                _excel_grafica_titular(ws, titulo, len(filas))

        borde = Border(*(Side(style='thin', color='dcdfe0'),) * 4)
        for ws in wb.worksheets:
            if ws.title == 'Portada':
                ws.sheet_properties.tabColor = VERDE_INSTITUCIONAL
                continue
            ws.sheet_properties.tabColor = VERDE_INSTITUCIONAL
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor=VERDE_INSTITUCIONAL)
                cell.alignment = Alignment(vertical='center')
            ws.freeze_panes = 'A2'
            ws.row_dimensions[1].height = 20
            for fila_idx in range(2, ws.max_row + 1):
                if fila_idx % 2 == 0:
                    for celda in ws[fila_idx]:
                        celda.fill = PatternFill('solid', fgColor='F2F6F4')
            for fila in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for celda in fila:
                    celda.border = borde
            for column in ws.columns:
                ws.column_dimensions[column[0].column_letter].width = min(max(len(str(c.value or '')) for c in column) + 3, 45)

        salida = BytesIO()
        wb.save(salida)
        contenido = salida.getvalue()
        response = HttpResponse(contenido, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="reporte_titular.xlsx"'
        return response


def _pdf_grafica_titular(seccion, filas):
    """Mismo criterio que _pdf_grafica de reportes.py (ROL1) — 'estados' y
    'tiempos' comparten exactamente la misma forma de datos ahí (nombre/
    cantidad/porcentaje y rango/cantidad), así que se reutiliza tal cual.
    'evolucion' y 'prioridad' sí necesitan su propia versión: aquí la
    evolución usa 'recibidos' (no 'registrados') y no existe un equivalente
    de barra por 'nombre' en reportes.py (solo 'unidad'/'rango')."""
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.legends import Legend
    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing

    if not filas:
        return None
    paleta = [_hex_color(f'#{h}') for h in PALETA_HEX]

    if seccion in ('estados', 'tiempos'):
        return _pdf_grafica(seccion, filas)

    if seccion == 'evolucion':
        d = Drawing(450, 185)
        chart = HorizontalLineChart()
        chart.x, chart.y = 35, 30
        chart.width, chart.height = 300, 135
        chart.data = [
            [f['recibidos'] for f in filas],
            [f['concluidos'] for f in filas],
            [f['vencidos'] for f in filas],
        ]
        chart.categoryAxis.categoryNames = [f['periodo'] for f in filas]
        chart.categoryAxis.labels.fontSize = 6.5
        chart.categoryAxis.labels.angle = 30
        chart.valueAxis.valueMin = 0
        if max((v for fila in chart.data for v in fila), default=0) == 0:
            chart.valueAxis.valueMax = 1
        etiquetas = ['Recibidos', 'Concluidos', 'Vencidos']
        for i in range(3):
            chart.lines[i].strokeColor = paleta[i]
            chart.lines[i].strokeWidth = 2
        d.add(chart)
        leyenda = Legend()
        leyenda.x, leyenda.y = 355, 145
        leyenda.fontSize = 7
        leyenda.colorNamePairs = list(zip(paleta[:3], etiquetas))
        d.add(leyenda)
        return d

    if seccion == 'prioridad':
        d = Drawing(450, 185)
        chart = VerticalBarChart()
        chart.x, chart.y = 40, 40
        chart.width, chart.height = 380, 130
        chart.data = [[f['cantidad'] for f in filas]]
        chart.categoryAxis.categoryNames = [f['nombre'] for f in filas]
        chart.categoryAxis.labels.fontSize = 7.5
        chart.valueAxis.valueMin = 0
        if max((v for fila in chart.data for v in fila), default=0) == 0:
            chart.valueAxis.valueMax = 1
        chart.bars[0].fillColor = paleta[0]
        d.add(chart)
        return d

    return None


class ReporteTitularExportarPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import HRFlowable, KeepTogether, Paragraph, SimpleDocTemplate, Spacer

        datos = _datos_titular(request)

        st_titulo = ParagraphStyle('rep_titulo', fontName='Helvetica-Bold', fontSize=19, leading=23, textColor=_hex_color('#' + VERDE_INSTITUCIONAL), spaceAfter=8)
        st_subtitulo = ParagraphStyle('rep_subtitulo', fontName='Helvetica', fontSize=9.5, leading=13, textColor=colors.HexColor('#555555'), spaceAfter=2)
        st_seccion = ParagraphStyle('rep_seccion', fontName='Helvetica-Bold', fontSize=12.5, textColor=colors.white, backColor=_hex_color('#' + VERDE_INSTITUCIONAL), leftIndent=0, spaceBefore=0)
        st_celda = ParagraphStyle('rep_celda', fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor('#1a2f28'), leading=9.5)
        st_celda_header = ParagraphStyle('rep_celda_header', parent=st_celda, fontName='Helvetica-Bold', textColor=colors.white)
        st_vacio = ParagraphStyle('rep_vacio', fontName='Helvetica-Oblique', fontSize=8.5, textColor=colors.HexColor('#7a7a7a'))

        ancho_util = letter[0] - 4 * cm  # 2cm de margen a cada lado, igual que el PDF de ROL1

        nombre_titular = resolver_nombre(request.user)
        filtros_aplicados = [
            f"{etiqueta}: {request.query_params.get(clave)}"
            for clave, etiqueta in [('estatus', 'Estado'), ('prioridad', 'Prioridad')]
            if request.query_params.get(clave)
        ]

        elementos = [
            Paragraph('Reporte de turnados — Titular', st_titulo),
            Paragraph('Inteligencia operativa — Sistema de Control de Solicitudes', st_subtitulo),
            Spacer(1, 6),
            HRFlowable(width='100%', thickness=1.4, color=_hex_color('#' + VERDE_INSTITUCIONAL), spaceAfter=8),
            Paragraph(f"<b>Titular:</b> {nombre_titular}", st_subtitulo),
            Paragraph(f"<b>Periodo:</b> {datos['periodo']['inicio']} a {datos['periodo']['fin']}", st_subtitulo),
        ]
        if filtros_aplicados:
            elementos.append(Paragraph(f"<b>Filtros aplicados:</b> {' · '.join(filtros_aplicados)}", st_subtitulo))
        elementos.append(Spacer(1, 16))

        contenido = {
            'resumen': datos['resumen'], 'evolucion': datos['evolucion'],
            'estados': datos['estados'], 'prioridad': datos['prioridad'],
            'tiempos': datos['tiempos'], 'detalle': datos['detalle'][:150],
        }
        for seccion in ('resumen', 'evolucion', 'estados', 'prioridad', 'tiempos', 'detalle'):
            filas = contenido[seccion]
            grafica = _pdf_grafica_titular(seccion, filas) if not isinstance(filas, dict) else None

            # Encabezado (+ gráfica, si tiene) siempre juntos, igual que en el
            # PDF de ROL1 — sin esto un encabezado puede quedar solo al fondo
            # de una página con su contenido empujado a la siguiente.
            bloque_encabezado = [Paragraph(f"&nbsp;{SECCIONES_TITULAR[seccion]}", st_seccion), Spacer(1, 10)]
            if grafica is not None:
                bloque_encabezado.extend([grafica, Spacer(1, 8)])
            elementos.append(KeepTogether(bloque_encabezado))

            if isinstance(filas, dict):
                tabla = [['Indicador', 'Valor']] + [[_etiqueta_campo(k), v] for k, v in filas.items()]
            elif filas:
                tabla = [[_etiqueta_campo(k) for k in filas[0]]] + [list(x.values()) for x in filas]
            else:
                tabla = None
            if tabla:
                elementos.append(_pdf_tabla_ejecutiva(tabla, st_celda, st_celda_header, ancho_util))
            else:
                elementos.append(Paragraph('Sin información para los filtros seleccionados.', st_vacio))
            elementos.append(Spacer(1, 18))

        salida = BytesIO()
        doc = SimpleDocTemplate(
            salida, pagesize=letter,
            leftMargin=2 * cm, rightMargin=2 * cm, topMargin=3.3 * cm, bottomMargin=4 * cm,
            title='Reporte de turnados — Titular', author='Sistema de Control de Solicitudes',
        )
        doc.build(
            elementos,
            onFirstPage=_pdf_membrete, onLaterPages=_pdf_membrete,
            canvasmaker=_pdf_construir_numbered_canvas(),
        )
        contenido_bytes = salida.getvalue()
        response = HttpResponse(contenido_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_titular.pdf"'
        return response
